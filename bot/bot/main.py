import os
import datetime
import time
import gspread
import openpyxl
from oauth2client.service_account import ServiceAccountCredentials
import pandas as pd
import telegram
import asyncio
import telegram.ext
import logging
import configparser


import traceback

# Create a ConfigParser object
config = configparser.ConfigParser()
config.read('config_test.ini')

# Fetch configuration from ini file
chat_id = int(config.get('Telegram', 'chat_id'))
BotToken = config.get('Telegram', 'BotToken').strip()
cust_sheet_url = config.get('GoogleSheets', 'customer_sheet').strip()
cust_sheet_name = config.get('GoogleSheets', 'cust_sheet_name').strip()
review_sheet_url = config.get('GoogleSheets', 'review_sheet').strip()
review_sheet_name = config.get('GoogleSheets', 'review_sheet_name').strip()
product_mapping_ini = config['ProductMapping']

# Setup directories
current_dir = str(os.getcwd())
logdir = os.path.join(current_dir,'DebugLogs')
get_time = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
logfile = os.path.join(logdir, f'Logs_{get_time}.log')
if not os.path.exists(logdir):
    os.makedirs(logdir)

# Set up logging module
logging.basicConfig(
    filename=logfile,  # Specify the log file name
    level=logging.DEBUG,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
)
log = logging.getLogger('SheetAutomationService')
# console_handler = logging.StreamHandler()
# console_handler.setLevel(logging.INFO)  # Adjust the log level if needed
# log.addHandler(console_handler)

# Set up CLI configuration
while True:
    global date_in, formatted_date
    try:
        date_in = input("Enter the date for which you want to fetch the data in (DD-MM-YYYY)\n Enter the date: ")
        parsed_date = datetime.datetime.strptime(date_in, "%d-%m-%Y")
        formatted_date = parsed_date.strftime("%Y-%m-%d")
        break
    except:
        print("please check the date you've entered is valid and as per the 'DD-MM-YYYY' format\nPlease try again....")
        print("~"*30)
        time.sleep(1)
        continue

# Set up GoogleSheet configuration
scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive',
         'https://www.googleapis.com/auth/drive.file', 'https://www.googleapis.com/auth/spreadsheets']
creds = ServiceAccountCredentials.from_json_keyfile_name('Cred.json', scope)
client = gspread.authorize(creds)
sheet = client.open_by_url(cust_sheet_url)
customer_sheet_obj = sheet.worksheet(cust_sheet_name)
review_sheet = client.open_by_url(review_sheet_url)
review_sheet_obj = review_sheet.worksheet(review_sheet_name)
review_file_xlsx = 'review_offline.xlsx'

# Initialize marking indexes for batch update
mark_review_cell = []
mark_customer_cell = []

# Functions
def map_product(product_mapping):
    product_mapping_dict = {}
    for key, value in product_mapping.items():
        product_mapping_dict[key.capitalize().strip()] = value.upper().strip()
    transformed_dict = {key: tuple(value.split(',')) for key, value in product_mapping_dict.items()}
    return transformed_dict

# Product mapping according to review sheet
product_mapping = map_product(product_mapping_ini)

def find_entry_in_excel(dataframe, search_value, column_to_search, excel_file_path):
    # Load the main Excel file
    workbook = openpyxl.load_workbook(excel_file_path)
    # Initialize a variable to store the cell index
    cell_index = None

    for index, row in dataframe.iterrows():
        if row[column_to_search] == search_value:
            # If a match is found, record the row and column indices
            cell_index = f"{row.name + 2}{column_to_search}"  # +2 to adjust for 0-based index and Excel's 1-based index
            break

    # Close the main Excel file
    workbook.close()
    if cell_index:
        # return cell index
        log.info(f"Cell index {cell_index}")
        return cell_index
    else:
        # log the error
        log.info(f"The entry '{search_value}' was not found in the Excel file.")
        return f"The entry '{search_value}' was not found in the Excel file."

def save_offline(sheet_obj, filename):
    # Get all values in the worksheet
    data = sheet_obj.get_all_values()
    df = pd.DataFrame(data[1:], columns=data[0])

    # Save the data to an Excel file (xlsx format)
    output_path = os.path.join(current_dir,filename)
    df.to_excel(str(output_path), index=False, engine='openpyxl')
    log.info(f"review file created offline {output_path}")
    return output_path

def read_worksheet(customersheet_obj):
    data = customersheet_obj.get_all_values()
    data_frame = pd.DataFrame(data[1:], columns=data[0])
    # get filtered dataframe based of MArking 'Done'
    filtered_data_frame = data_frame[data_frame['Marked'] != 'Done']
    log.info(f"Filtered Customer Data for 'Done'")
    return filtered_data_frame

def filter_data_on_date(df, date):
    df['Delivery date for reviews'] = pd.to_datetime(df['Delivery date for reviews'])
    # Define the target delivery date
    target_delivery_date = pd.to_datetime(str(date))
    # yyyy-mm-dd
    # Filter the DataFrame based on the target delivery date
    filtered_df = df[df['Delivery date for reviews'] == target_delivery_date]
    if filtered_df.empty:
        # No customer found or left to mark
        log.info(f"No Customer data for given date {date}")
        return None
    else:
        # return filtered dataframe based on date
        log.info(f"Filtered customer data for given date {date}")
        return filtered_df

# ------------------------------------------------ Not actively used  -----------------------------------------------
def fetch_review(prod, reviewsheet_obj):
    '''
    Not being used in the code, just for future extention based on total online updation on google sheet
    '''
    product=prod.strip()
    print(f"fectching for {product}")
    # Find the corresponding column and marking column for the product
    column, marking_column = product_mapping.get(product)
    data = reviewsheet_obj.get_all_values()
    rows = 0
    for row in data:
        if row and row[0] != "":
            rows +=1

    for i in range(2,rows):
        # Fetch the review and marking from the Google Sheet
        marking = reviewsheet_obj.acell(f"{marking_column}{i}").value
        # Check if the marking is not "Done" before using the review
        if marking != "Done":
            # print(f"got the value for {column}{i}")
            review = reviewsheet_obj.acell(f"{column}{i}").value
            log.info(f"Unique review feteched for {product} online")
            return review
        else:
            continue
# ------------------------------------------------------------------------------------------------

def fetch_review_offline(prod, excel_file_path):
    product = prod.capitalize().strip()
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(excel_file_path)
    # Get the desired worksheet by name or index
    worksheet = workbook["Sheet1"]  # Replace "Sheet1" with the actual sheet name
    # Find the corresponding column and marking column for the product
    column, marking_column = product_mapping.get(product)
    # Get the maximum row count in the worksheet
    rows = worksheet.max_row
    for i in range(2, rows + 1):  # Openpyxl uses 1-based indexing
        # Fetch the review and marking from the Excel file
        marking = worksheet[f"{marking_column}{i}"].value
        # Check if the marking is not "Done" before using the review
        if marking != "Done" and f"{marking_column}{i}" not in mark_review_cell:
            review = worksheet[f"{column}{i}"].value
            log.info(f"Unique review feteched for {product} from {review_file_xlsx}")
            # Check if the review is not empty
            if review is not None:
                # is_marked = mark_cell_done(marking_column, i, worksheet)
                mark_review_cell.append(f"{marking_column}{i}")
                worksheet[f"{marking_column}{i}"].value = "Done"
                workbook.save(excel_file_path)
                workbook.close()
                log.info(f"review fetched for {product}")
                # return review for given product
                return review
            else:
                log.error(f"No unique review left in the sheet for product: {product} or product not found in the review sheet.\n Make sure colums are mapped correctly in config file for the product: {product}")
                print(f"No unique review left in the sheet for product: {product} or product not found in the review sheet.\n Make sure colums are mapped correctly in config file for the product: {product}")
                return f"No unique review left in the sheet for product: {product} or product not found in the review sheet.\n Make sure colums are mapped correctly in config file for the product: {product}"
        else:
            # Continue looping if previous review is already marked
            continue

async def send(chat, msg):
    await telegram.Bot(BotToken).sendMessage(chat_id=chat, text=msg)

counter = 0
def send_review(customer_name, product_review_dict, mobile_num, point_of_contact, date_, order_id):
    global counter
    # Function to send the review to the given chat id along with the formated review
    data1 = customer_name
    data2 = product_review_dict
    final_data = dict(customer_name = data1, procuct_list = data2, point_of_contact=point_of_contact, mobile_num=mobile_num, Delivery_date=date_, order_id=order_id)
    # Format the dictionary to message templet
    formatted_data_f, product_revs = format_data(final_data)
    # Send the message
    # print(f"counter: {counter}")
    counter += 1
    log.info(f"counter {counter}")
    if counter == 19:
        print("~" * 100)
        print("Cool down period of 30 Sec...")
        print("~" * 100)
        time.sleep(30)
        log.info("Cool down period of 30 Sec...")
        counter = 0
        # print(f"count = {counter}", flush=True)
        log.info(f"counter {counter}")

    loop = asyncio.get_event_loop()
    loop.run_until_complete(send(chat=chat_id, msg=str(formatted_data_f)))
    for products in product_revs:
        counter += 1
        log.info(f"counter {counter}")
        if counter == 19:
            print("~" * 100)
            print("Cool down period of 30 Sec...")
            print("~" * 100)
            time.sleep(30)
            log.info("Cool down period of 30 Sec...")
            counter = 0
            # print(f"count = {counter}", flush=True)
            log.info(f"counter {counter}")
        loop = asyncio.get_event_loop()
        loop.run_until_complete(send(chat=chat_id, msg=str(products)))
    # asyncio.run(send(chat=chat_id, msg=str(formatted_data_f)))
    log.info(f"message sent to chat_id: {chat_id} for customer review")


def fetch_all_product_review(filtered_df, reviewsheet_obj, date_, customersheet_obj):

    # Function to fetch all reviews
    filterd_df_dict = filtered_df
    # Download the review sheet
    offline_review_file = save_offline(reviewsheet_obj, filename=review_file_xlsx)
    # Iterate over the sheet for each product in filtered_df_dict
    for index in range(len(filterd_df_dict)):
        try:
            # Fetch the various product details
            item = filterd_df_dict[index]
            item = dict(item)
            customer_name = item["Customer name"]
            mobile_no = item["Mobile number"]
            point_of_contact = item["Point of contact"]
            order_id = item["Order id"]
            product_list = item["Product"].split(",") if len(item["Product"]) > 1 else [item["Product"]]
            print(f"customer: {customer_name} || products: {product_list}")
            review_dict = {}
            for item in product_list:
                prod_name = str(item.capitalize().strip())
                # Fetches the product review
                review_rec = fetch_review_offline(prod=prod_name, excel_file_path=offline_review_file)
                review_dict[prod_name] = review_rec
                # Map the customer name to their review dictionary
            # Send the review
            send_review(customer_name=customer_name, product_review_dict=review_dict, mobile_num=mobile_no, point_of_contact=point_of_contact, date_=date_, order_id=order_id)
            time.sleep(0.5)
            # Get the customer index based on unique order id
            cust_index = get_customer_row_index(customersheet_obj=customersheet_obj, order_id=order_id)
            if cust_index is not None:
                mark_customer_cell.append(f"K{cust_index}")
            else:
                # print(f"customer index not found for {customer_name}")
                log.error(f"customer index not found for {customer_name}")
        except Exception as e:
            # traceback.print_exc()
            log.error(f"func: [fetch_all_product_review] error:{e} caused unexpected exit")
            print("#"*100)
            print(f"Error: {e}")
            print("#" * 100)
            continue

def format_data(input_dict):
    formatted_data = []
    product_rev = []
    # Extract common information
    customer_name = input_dict['customer_name']
    customer_contact = input_dict['mobile_num']
    point_of_contact = input_dict['point_of_contact']
    delivery_date = input_dict['Delivery_date']
    order_id = input_dict['order_id']

    # Create the header section
    header = f"For Delivery date {delivery_date}\n\n"
    header += f"Customer name = {customer_name}\n"
    header += f"Customer contact = {customer_contact}\n"
    header += f"Point of contact = {point_of_contact}\n"
    header += f"Order id = {order_id}\n"
    header += "-" * 50 + "\n"

    formatted_data.append(header)

    # Loop over the product_list
    product_list = input_dict['procuct_list']
    for product, review in product_list.items():
        product_review = f"Product: {product}\n\n{review}\n"
        product_rev.append(product_review)
        # formatted_data.append(product_review)

    return '\n'.join(formatted_data), product_rev

def get_customer_row_index(customersheet_obj, order_id):
    # Implimented on Order ID instead of customer name, if customer repeated then order id will be unique each time
    # Load the Excel file into a Pandas DataFrame
    data = customersheet_obj.get_all_values()
    df = pd.DataFrame(data[1:], columns=data[0])

# Check if the "Order id" column exists in the DataFrame
    if 'Order id' in df.columns:
        # Search for the customer name in the "Order id" column
        matching_rows = df[df['Order id'] == order_id]
        # Check if any matching rows were found
        if not matching_rows.empty:
            # Get the row index of the first matching row
            row_index = matching_rows.index[0]
            return row_index+2
    return None

def convert_to_update_requests(cell_indexes):
    try:
        # Convert cell indexes to JSON objects
        return [{'range': cell, 'values':[["Done"]]} for cell in cell_indexes]
    except Exception as e:
        # traceback.print_exc()
        # print(f"func:[convert_to_update_requests] error: {e} caused unexpected exit ")
        log.error(f"func:[convert_to_update_requests] error: {e} caused unexpected exit ")

def batch_update_cells(sheet, cell_updates, sheet_name):
    try:
        # Batch update according to JSON object passed
        sheet.batch_update(cell_updates)
        log.info(f"Updated data for {sheet_name}")
        # print(f"updated: {sheet_name}")
    except Exception as e:
        # traceback.print_exc()
        # print(f"func:[batch_update_cells] error: {e} caused unexpected exit ")
        log.error(f"func:[batch_update_cells] error: {e} caused unexpected exit ")

# --------------------------------------------------------------------------------------------------------------------

try:
    all_df = read_worksheet(customersheet_obj=customer_sheet_obj)
    filtered_df = filter_data_on_date(all_df, str(formatted_date))
    if filtered_df is None:
        loop = asyncio.get_event_loop()
        loop.run_until_complete(send(chat=chat_id, msg=f"No Customer left for the given date: {date_in}\n or Date does not exists in the sheet"))
        # asyncio.run()

        print(f"No Customer left for the given date: {date_in} or Date does not exists in the sheet")
        log.info(f"No Customer left for the given date: {date_in} or Date does not exists in the sheet")
        exit()
    filtered_df_dict = filtered_df.to_dict(orient='records')
    fetch_all_product_review(filtered_df_dict, reviewsheet_obj=review_sheet_obj, date_=str(formatted_date), customersheet_obj=customer_sheet_obj)
    log.info(f"review_cell markings {mark_review_cell}")
    log.info(f"review_cell markings {mark_customer_cell}")

    print("~" * 100)
    print("~"*30, "All data fetched","~"*30)
    print("~" * 100)
    print("~"*30, "Review Sheet indexes","~"*30)
    print(mark_review_cell)
    print("~" * 100)
    print("~" * 30, "Customer Sheet indexes", "~" * 30)
    print(mark_customer_cell)
    cell_indexes_rev= convert_to_update_requests(mark_review_cell)
    if len(cell_indexes_rev) == 0:
        print(f"empty cell_indexes for {cell_indexes_rev}")
    cell_indexes_cust = convert_to_update_requests(mark_customer_cell)
    if len(cell_indexes_cust) == 0:
        print(f"empty cell_indexes for {cell_indexes_cust}")
    batch_update_cells(sheet=review_sheet_obj, cell_updates=cell_indexes_rev, sheet_name="Reviews Sheet")
    batch_update_cells(sheet=customer_sheet_obj, cell_updates=cell_indexes_cust, sheet_name="Customer Data Sheet")
    os.remove(os.path.join(os.getcwd(), review_file_xlsx))
except Exception as e:
    # traceback.print_exc()
    print("Update following cells in respective GoogleSheets if code exited unexpectedly")
    print(f"Update review sheet at index {mark_review_cell}")
    print(f"Update review sheet at index {mark_customer_cell}")
    os.remove(os.path.join(os.getcwd(), review_file_xlsx))
